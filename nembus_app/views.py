# nembus_app/views.py
import locale
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .models import (
    Cliente, Camion, ReporteVenta, PerfilTrabajador, Traspaso,
    PuntoDeVenta, Bomba, Turno, ReporteTurno, LecturaBomba,
    RegistroVentaIndividualBomba # Importar nuevo modelo
)
from decimal import Decimal
from django.contrib import messages
from django.utils import timezone # Asegúrate que timezone esté importado
from datetime import timedelta, datetime # Asegúrate que datetime y timedelta estén importados
from django.db.models import Sum, Count, F
from django.db.models.functions import TruncDay, TruncHour
import json
import csv
from django.http import HttpResponse
# Imports para nuevos forms y lógica de turno
from .forms import IniciarTurnoForm, VentaIndividualFormSet # Importar nuevos forms
from django.forms import inlineformset_factory
from django.db import transaction # Para guardar formsets atomicamente
# Imports para Excel
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# Imports para LogEntry
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE, DELETION # Importar LogEntry y constantes
from django.contrib.contenttypes.models import ContentType # Importar ContentType
from openpyxl.utils import get_column_letter

# --- VISTAS DE AUTENTICACIÓN Y AUXILIARES ---
def login_usuario(request):
    error_message = None
    if request.method == 'POST':
        user = authenticate(request, username=request.POST.get('username'), password=request.POST.get('password'))
        if user is not None:
            login(request, user)
            # Redirigir según el tipo de usuario
            try:
                # Usar get_or_create para manejar usuarios sin perfil inicial
                perfil, created = PerfilTrabajador.objects.get_or_create(usuario=user)
                if perfil.punto_de_venta_asignado:
                    # Es bombero
                    return redirect('nembus_app:dashboard_trabajador')
                elif user.is_superuser:
                    # Es admin/gerente
                    return redirect('nembus_app:dashboard_gerente_redirect')
                else:
                    # Es chofer (sin punto de venta asignado)
                    return redirect('nembus_app:dashboard_trabajador')
            except Exception: # Captura genérica por si algo falla al obtener/crear perfil
                 # Por defecto, si falla, intentar ir al dashboard trabajador (o login si prefieres)
                 if user.is_superuser:
                     return redirect('nembus_app:dashboard_gerente_redirect')
                 else:
                    return redirect('nembus_app:dashboard_trabajador')
        else:
            error_message = "Usuario o contraseña incorrectos."
    # Si es GET o fallo el login
    return render(request, 'nembus_app/login.html', {'error': error_message})

def logout_usuario(request):
    logout(request)
    messages.info(request, "Has cerrado sesión.") # Mensaje opcional
    return redirect('nembus_app:login')

@login_required
def reporte_exito(request): # Para éxito de ReporteVenta (camiones)
    # Considera hacerlo más genérico o tener páginas de éxito específicas
    return render(request, 'nembus_app/reporte_exito.html')

# --- VISTAS PARA EL TRABAJADOR (CHOFER Y BOMBERO) ---
@login_required
def dashboard_trabajador(request):
    try:
        # Usar get_or_create para asegurar que el perfil exista
        perfil, created = PerfilTrabajador.objects.get_or_create(usuario=request.user)
    except Exception as e:
         messages.error(request, f"Error al cargar tu perfil: {e}")
         logout(request) # Forzar logout si no se puede cargar el perfil
         return redirect('nembus_app:login')

    context = {}
    if perfil.punto_de_venta_asignado:
        context['es_bombero'] = True
        context['punto_de_venta'] = perfil.punto_de_venta_asignado
        # Verificar si tiene un turno activo
        turno_abierto = ReporteTurno.objects.filter(trabajador=request.user, esta_abierto=True).first()
        context['turno_activo'] = turno_abierto # Será None si no hay turno activo
    else: # Es Chofer
        context['es_bombero'] = False
        context['tiene_permiso_recarga'] = perfil.puede_recargar_combustible
        context['puede_hacer_traspasos'] = perfil.puede_hacer_traspasos

    return render(request, 'nembus_app/dashboard_trabajador.html', context)

# --- VISTAS PARA CHOFERES (OPERACIONES CON CAMIONES) ---
@login_required
def crear_reporte_venta(request): # Ventas desde CAMIÓN
    try:
        perfil = request.user.perfiltrabajador
        if perfil.punto_de_venta_asignado: # No permitir a bomberos
             messages.error(request, "Acción no disponible para tu perfil.")
             return redirect('nembus_app:dashboard_trabajador')
    except PerfilTrabajador.DoesNotExist:
        messages.error(request, "Perfil no configurado.")
        return redirect('nembus_app:dashboard_trabajador')

    if request.method == 'POST':
        try:
            # Obtener datos del POST de forma segura
            cliente_id = request.POST.get('cliente')
            camion_id = request.POST.get('camion')
            litros_str = request.POST.get('litros')

            if not cliente_id or not camion_id or not litros_str:
                raise ValueError("Faltan datos obligatorios.")

            cliente = perfil.clientes_asignados.get(id=cliente_id)
            camion = perfil.camiones_asignados.get(id=camion_id)
            litros_vendidos = Decimal(litros_str)

            if litros_vendidos <= 0:
                 raise ValueError("Los litros vendidos deben ser positivos.")

            if camion.litros_actuales >= litros_vendidos:
                with transaction.atomic(): # Asegurar consistencia
                    # Actualizar litros del camión ANTES de guardar el reporte
                    camion.litros_actuales -= litros_vendidos
                    camion.save(update_fields=['litros_actuales'])

                    monto_combustible = litros_vendidos * cliente.precio_litro_clp
                    costo_flete = cliente.costo_flete_clp
                    reporte = ReporteVenta(
                        trabajador=request.user, cliente=cliente, camion=camion, litros_vendidos=litros_vendidos,
                        monto_combustible_clp=monto_combustible, costo_flete_clp=costo_flete,
                        monto_total_clp=monto_combustible + costo_flete
                    )
                    if 'foto' in request.FILES:
                        reporte.foto_evidencia = request.FILES['foto']
                    reporte.save() # Guardar el reporte

                    # --- REGISTRAR ACCIÓN EN ADMIN LOG (VERIFICAR ESTO) ---
                    try:
                       LogEntry.objects.log_action(
                           user_id=request.user.id,
                           content_type_id=ContentType.objects.get_for_model(reporte).pk,
                           object_id=reporte.pk,
                           object_repr=str(reporte), # Representación textual del objeto
                           action_flag=ADDITION, # Indicar que fue una adición (creación)
                           change_message="Venta registrada desde formulario web." # Mensaje opcional
                       )
                       print(f"LogEntry registrado para ReporteVenta ID: {reporte.pk}") # Añadir confirmación
                    except Exception as log_error:
                       # **ACCIÓN:** Si las ventas de camión no aparecen en 'Recent Actions',
                       # busca este mensaje en los logs de tu servidor Django.
                       print(f"Error al registrar LogEntry para ReporteVenta: {log_error}")
                    # --- FIN REGISTRO ACCIÓN ---

                messages.success(request, "¡Venta de camión guardada con éxito!")
                return redirect('nembus_app:dashboard_trabajador') # Asegurar namespace
            else:
                messages.error(request, f"Error: No hay suficientes litros en el camión ({camion.patente}). Disponibles: {camion.litros_actuales} L.")
        except Cliente.DoesNotExist:
             messages.error(request, "Cliente seleccionado no válido.")
        except Camion.DoesNotExist:
             messages.error(request, "Camión seleccionado no válido.")
        except (ValueError, TypeError) as e:
             messages.error(request, f"Dato inválido: {e}")
        except Exception as e:
             messages.error(request, f"Ocurrió un error inesperado: {e}")

    # Para GET o si hubo error en POST
    context = {'clientes': perfil.clientes_asignados.all(), 'camiones': perfil.camiones_asignados.all()}
    return render(request, 'nembus_app/crear_reporte.html', context)


@login_required
def crear_recarga(request): # Recarga de CAMIÓN
    try:
        perfil = request.user.perfiltrabajador
        if not perfil.puede_recargar_combustible or perfil.punto_de_venta_asignado: # Solo para choferes con permiso
            messages.error(request, "Acción no permitida.")
            return redirect('nembus_app:dashboard_trabajador')
    except PerfilTrabajador.DoesNotExist:
         messages.error(request, "Perfil no configurado.")
         return redirect('nembus_app:dashboard_trabajador')

    camiones_asignados = perfil.camiones_asignados.all() # Camiones que puede recargar

    if request.method == 'POST':
        try:
            camion_id = request.POST.get('camion')
            litros_str = request.POST.get('litros')

            if not camion_id or not litros_str:
                 raise ValueError("Faltan datos obligatorios.")

            camion = camiones_asignados.get(id=camion_id)
            litros_a_recargar = Decimal(litros_str)

            if litros_a_recargar <= 0:
                 messages.error(request, "La cantidad a recargar debe ser positiva.")
            # Validar capacidad
            elif camion.litros_actuales + litros_a_recargar > camion.capacidad_total:
                messages.error(request, f"Error: La recarga excede la capacidad de {camion.patente}. Máximo a añadir: {camion.capacidad_total - camion.litros_actuales} L.")
            else:
                with transaction.atomic(): # Usar transacción por si se añade LogEntry
                    camion.litros_actuales += litros_a_recargar
                    camion.save(update_fields=['litros_actuales'])
                    messages.success(request, f"¡Recarga de {litros_a_recargar}L guardada con éxito para {camion.patente}!")

                    # --- OPCIONAL: Registrar Recarga en LogEntry ---
                    try:
                       LogEntry.objects.log_action(
                           user_id=request.user.id,
                           content_type_id=ContentType.objects.get_for_model(camion).pk,
                           object_id=camion.pk,
                           object_repr=str(camion),
                           action_flag=CHANGE, # Es un cambio en el camión
                           change_message=f"Recarga de {litros_a_recargar}L registrada desde formulario web."
                       )
                       print(f"LogEntry registrado para Recarga Camión ID: {camion.pk}")
                    except Exception as log_error:
                       print(f"Error al registrar LogEntry para Recarga Camión: {log_error}")
                    # --- FIN OPCIONAL ---

                return redirect('nembus_app:dashboard_trabajador')
        except Camion.DoesNotExist:
            messages.error(request, "Camión seleccionado no válido.")
        except (ValueError, TypeError) as e:
            messages.error(request, f"Dato inválido: {e}")
        except Exception as e:
            messages.error(request, f"Ocurrió un error inesperado: {e}")

    # Para GET o si hubo error en POST
    context = {'camiones': camiones_asignados}
    return render(request, 'nembus_app/crear_recarga.html', context)


@login_required
def crear_traspaso(request): # Traspaso entre CAMIONES
    try:
        perfil = request.user.perfiltrabajador
        if not perfil.puede_hacer_traspasos or perfil.punto_de_venta_asignado: # Solo choferes con permiso
            messages.error(request, "Acción no permitida.")
            return redirect('nembus_app:dashboard_trabajador')
    except PerfilTrabajador.DoesNotExist:
        messages.error(request, "Perfil no configurado.")
        return redirect('nembus_app:dashboard_trabajador')

    camiones_para_traspaso = perfil.camiones_traspaso.all() # Camiones habilitados para traspaso
    if not camiones_para_traspaso.exists():
         messages.warning(request, "No tienes camiones asignados para realizar traspasos.")
         # Considerar mostrar mensaje en la plantilla en lugar de redirigir

    if request.method == 'POST':
        try:
            origen_id = request.POST.get('camion_origen')
            destino_id = request.POST.get('camion_destino')
            litros_str = request.POST.get('litros')

            if not origen_id or not destino_id or not litros_str:
                 raise ValueError("Faltan datos obligatorios.")

            camion_origen = camiones_para_traspaso.get(id=origen_id)
            camion_destino = camiones_para_traspaso.get(id=destino_id)
            litros_a_traspasar = Decimal(litros_str)

            if litros_a_traspasar <= 0:
                 messages.error(request, "La cantidad a traspasar debe ser positiva.")
            elif camion_origen == camion_destino:
                messages.error(request, "Error: El camión de origen y destino no pueden ser el mismo.")
            elif camion_origen.litros_actuales < litros_a_traspasar:
                messages.error(request, f"Error: No hay suficientes litros en {camion_origen.patente}. Disponibles: {camion_origen.litros_actuales} L.")
            # Validar capacidad destino
            elif camion_destino.litros_actuales + litros_a_traspasar > camion_destino.capacidad_total:
                messages.error(request, f"Error: El traspaso excede la capacidad de {camion_destino.patente}.")
            else:
                with transaction.atomic(): # Asegurar atomicidad
                    # Actualizar ambos camiones
                    camion_origen.litros_actuales -= litros_a_traspasar
                    camion_destino.litros_actuales += litros_a_traspasar
                    camion_origen.save(update_fields=['litros_actuales'])
                    camion_destino.save(update_fields=['litros_actuales'])
                    # Crear el registro del traspaso
                    traspaso_obj = Traspaso.objects.create( # Guardar en variable
                        trabajador=request.user, camion_origen=camion_origen,
                        camion_destino=camion_destino, litros=litros_a_traspasar
                    )
                    # --- REGISTRAR ACCIÓN EN ADMIN LOG (YA PRESENTE, ASEGURAR QUE FUNCIONE) ---
                    try:
                       LogEntry.objects.log_action(
                           user_id=request.user.id,
                           content_type_id=ContentType.objects.get_for_model(traspaso_obj).pk,
                           object_id=traspaso_obj.pk,
                           object_repr=str(traspaso_obj),
                           action_flag=ADDITION,
                           change_message="Traspaso registrado desde formulario web."
                       )
                       print(f"LogEntry registrado para Traspaso ID: {traspaso_obj.pk}") # Añadir confirmación
                    except Exception as log_error:
                       # **ACCIÓN:** Revisa logs si esta acción no aparece.
                       print(f"Error al registrar LogEntry para Traspaso: {log_error}")
                    # --- FIN REGISTRO ACCIÓN ---

                messages.success(request, f"¡Traspaso de {litros_a_traspasar}L guardado con éxito!")
                return redirect('nembus_app:dashboard_trabajador')
        except Camion.DoesNotExist:
            messages.error(request, "Error: Camión no válido o no permitido para traspasos.")
        except (ValueError, TypeError) as e:
             messages.error(request, f"Dato inválido: {e}")
        except Exception as e:
            messages.error(request, f"Ocurrió un error inesperado: {e}")

    # Para GET o si hubo error en POST
    context = {'camiones': camiones_para_traspaso}
    return render(request, 'nembus_app/crear_traspaso.html', context)

# --- VISTAS PARA BOMBEROS (NUEVO FLUJO DE TURNOS) ---

@login_required
def iniciar_turno(request):
    # Obtener perfil y punto de venta del bombero
    try:
        perfil = request.user.perfiltrabajador
        punto_venta = perfil.punto_de_venta_asignado
        if not punto_venta: # Asegurar que es bombero
            messages.error(request, "Acción no permitida para tu perfil.")
            return redirect('nembus_app:dashboard_trabajador')
    except PerfilTrabajador.DoesNotExist:
        messages.error(request, "Tu perfil de trabajador no está configurado.")
        return redirect('nembus_app:login')

    # Verificar si ya existe un turno abierto para este trabajador
    turno_abierto = ReporteTurno.objects.filter(trabajador=request.user, esta_abierto=True).first()
    if turno_abierto:
        messages.info(request, f"Ya tienes un turno activo iniciado el {turno_abierto.fecha_inicio.strftime('%d/%m %H:%M')}.")
        # Redirigir a la vista para gestionar el turno existente
        return redirect('nembus_app:gestionar_turno', reporte_id=turno_abierto.id)

    # Obtener las bombas del punto de venta para mostrar los campos de contador
    bombas = Bomba.objects.filter(punto_de_venta=punto_venta).order_by('nombre')
    if not bombas.exists():
         messages.warning(request, "No hay bombas configuradas en tu punto de venta. No se puede iniciar turno.")
         return redirect('nembus_app:dashboard_trabajador')

    if request.method == 'POST':
        # Pasar el punto_venta al formulario para que filtre los turnos y genere los campos correctos
        form = IniciarTurnoForm(request.POST, punto_venta=punto_venta)
        if form.is_valid():
            try:
                with transaction.atomic(): # Usar transacción para crear reporte y lecturas juntos
                    # Crear el ReporteTurno principal
                    nuevo_reporte = ReporteTurno.objects.create(
                        trabajador=request.user,
                        turno=form.cleaned_data['turno'],
                        esta_abierto=True # Marcarlo como activo
                        # fecha_inicio se establece por defecto (default=timezone.now)
                    )
                    # Crear una LecturaBomba para cada bomba con su contador inicial
                    lecturas_creadas = []
                    for bomba in bombas:
                        contador_field_name = f'contador_inicial_{bomba.id}'
                        contador_inicial_val = form.cleaned_data.get(contador_field_name)
                        if contador_inicial_val is not None:
                            lectura = LecturaBomba.objects.create(
                                reporte_turno=nuevo_reporte,
                                bomba=bomba,
                                contador_inicial=contador_inicial_val
                            )
                            lecturas_creadas.append(lectura)
                        else:
                             # Si alguna bomba no tuvo contador, abortar transacción
                             raise ValueError(f"Falta contador inicial para bomba {bomba.nombre}")

                    # Verificar si se creó al menos una lectura
                    if not lecturas_creadas:
                        raise ValueError("No se pudo iniciar el turno, no se registraron lecturas iniciales.")

                    # --- REGISTRAR ACCIÓN INICIO TURNO (YA PRESENTE, ASEGURAR QUE FUNCIONE) ---
                    try:
                       LogEntry.objects.log_action(
                           user_id=request.user.id,
                           content_type_id=ContentType.objects.get_for_model(nuevo_reporte).pk,
                           object_id=nuevo_reporte.pk,
                           object_repr=str(nuevo_reporte),
                           action_flag=ADDITION,
                           change_message="Turno iniciado desde formulario web."
                       )
                       print(f"LogEntry registrado para Inicio de Turno ID: {nuevo_reporte.pk}") # Confirmación
                    except Exception as log_error:
                       # **ACCIÓN:** Revisa logs si esta acción no aparece.
                       print(f"Error al registrar LogEntry para inicio de ReporteTurno: {log_error}")
                    # --- FIN REGISTRO ACCIÓN ---

                messages.success(request, f"Turno iniciado correctamente (ID: {nuevo_reporte.id}). Ahora puedes registrar las ventas individuales.")
                # Redirigir a la vista de gestión del turno recién creado
                return redirect('nembus_app:gestionar_turno', reporte_id=nuevo_reporte.id)

            except ValueError as ve: # Capturar error si falta contador
                 messages.error(request, str(ve))
            except Exception as e:
                # Capturar otros posibles errores (ej. de base de datos)
                messages.error(request, f"Error inesperado al iniciar el turno: {e}")
                # El formulario se volverá a mostrar con los datos ingresados
    else: # Método GET
        # Crear un formulario vacío, pasándole el punto_venta para que se configure correctamente
        form = IniciarTurnoForm(punto_venta=punto_venta)

    context = {
        'form': form,
        'punto_de_venta': punto_venta,
        'bombas': bombas # Necesario para renderizar los campos en la plantilla
    }
    return render(request, 'nembus_app/iniciar_turno.html', context)


# --- VISTA GESTIONAR_TURNO CON CORRECCIÓN DE GUARDADO Y LOGGING OPCIONAL ---
@login_required
def gestionar_turno(request, reporte_id):
    reporte = get_object_or_404(ReporteTurno, id=reporte_id, trabajador=request.user, esta_abierto=True)
    lecturas = reporte.lecturas.select_related('bomba').order_by('bomba__nombre').all()
    formsets_dict = {} # Para pasar formsets con errores al contexto

    if request.method == 'POST':
        print("\n--- Depurando gestionar_turno POST ---")
        finalizando_turno = 'finalizar_turno' in request.POST
        print(f"Finalizando turno: {finalizando_turno}")
        formsets_validos = True
        formsets_procesados = {}

        # 1. Instanciar y Validar TODOS los formsets
        for lectura in lecturas:
            prefix = f'ventas_{lectura.id}'
            # Pasar request.POST y la instancia de LecturaBomba al formset
            formset = VentaIndividualFormSet(request.POST, instance=lectura, prefix=prefix)
            formsets_procesados[lectura.id] = formset
            print(f"Instanciando formset para lectura {lectura.id} (Bomba: {lectura.bomba.nombre})")
            if not formset.is_valid():
                formsets_validos = False
                print(f"!!! Formset para lectura {lectura.id} NO es válido. Errores:")
                # Imprimir errores específicos de cada formulario dentro del formset
                for form_idx, form_errors in enumerate(formset.errors):
                    if form_errors:
                         print(f"  - Form {form_idx}: {form_errors} | Datos: {formset.forms[form_idx].data}")
                # Imprimir errores no ligados a un formulario específico
                if formset.non_form_errors():
                    print(f"  - Non-form errors: {formset.non_form_errors()}")
                messages.error(request, f"Hay errores en los datos de venta para la bomba '{lectura.bomba.nombre}'. Por favor, corrígelos.")

        print(f"Todos los formsets son válidos: {formsets_validos}")
        # 2. Si TODOS los formsets son válidos
        if formsets_validos:
            try:
                print("Intentando guardar dentro de transaction.atomic...")
                with transaction.atomic():
                    num_ventas_guardadas_total = 0
                    num_ventas_borradas_total = 0

                    # Iterar sobre los formsets validados
                    for lectura_id, formset in formsets_procesados.items():
                        print(f"  Procesando formset para lectura {lectura_id}")
                        lectura_obj = LecturaBomba.objects.select_related('bomba').get(id=lectura_id)
                        num_saved_this_fs = 0
                        num_deleted_this_fs = 0

                        # --- LÓGICA DE GUARDADO Y LOGGING ---
                        for form in formset:
                            # Verificar borrado
                            if form.cleaned_data.get('DELETE', False):
                                if form.instance.pk: # Solo si ya existe en la BD
                                    venta_pk_borrada = form.instance.pk # Guardar ID para log
                                    venta_repr_borrada = str(form.instance) # Guardar repr para log
                                    print(f"    Intentando borrar venta ID {venta_pk_borrada}")
                                    form.instance.delete()
                                    num_deleted_this_fs += 1
                                    # --- OPCIONAL: Registrar borrado en LogEntry ---
                                    try:
                                        LogEntry.objects.log_action(
                                            user_id=request.user.id,
                                            content_type_id=ContentType.objects.get_for_model(RegistroVentaIndividualBomba).pk,
                                            object_id=venta_pk_borrada, # Usar ID guardado
                                            object_repr=venta_repr_borrada, # Usar repr guardado
                                            action_flag=DELETION,
                                            change_message=f"Venta de bomba borrada desde gestión de turno (Lectura ID: {lectura_id})."
                                        )
                                        print(f"LogEntry (DELETE) registrado para Venta Bomba ID: {venta_pk_borrada}") # Confirmación
                                    except Exception as log_error:
                                        print(f"Error al registrar LogEntry (DELETE Venta Bomba): {log_error}")
                                    # --- FIN OPCIONAL ---

                            # Si no se borra y ha cambiado (nuevo o modificado)
                            # form.has_changed() es True para forms nuevos con datos y forms modificados
                            elif form.has_changed():
                                print(f"    Formulario {form.prefix} ha cambiado. Guardando...")
                                instance = form.save(commit=False) # Obtener la instancia SIN guardar en DB todavía

                                # *** LÍNEA CRÍTICA ASEGURADA ***
                                # Asegurar que la instancia (nueva o existente) sepa a qué LecturaBomba pertenece
                                instance.lectura_bomba = lectura_obj
                                # ******************************

                                # Asignar el precio de la bomba (esto ya parece estar bien en tu código)
                                instance.precio_litro_venta = lectura_obj.bomba.precio_litro_clp

                                # Determinar si es una adición o cambio para el LogEntry
                                action_flag_log = ADDITION if not instance.pk else CHANGE

                                # El método save() del modelo calculará 'ingreso_registro'
                                print(f"    Intentando guardar/actualizar: Máq={instance.numero_maquina}, Litros={instance.litros_vendidos}")
                                instance.save() # Guardar la instancia individual en la base de datos
                                num_saved_this_fs += 1
                                print(f"    Venta ID {instance.id} guardada/actualizada.")

                                # --- OPCIONAL: Registrar adición/cambio en LogEntry ---
                                try:
                                    LogEntry.objects.log_action(
                                        user_id=request.user.id,
                                        content_type_id=ContentType.objects.get_for_model(instance).pk,
                                        object_id=instance.pk,
                                        object_repr=str(instance),
                                        action_flag=action_flag_log,
                                        change_message=f"Venta de bomba {'creada' if action_flag_log == ADDITION else 'modificada'} desde gestión de turno (Lectura ID: {lectura_id})."
                                    )
                                    print(f"LogEntry ({'ADD' if action_flag_log == ADDITION else 'CHANGE'}) registrado para Venta Bomba ID: {instance.pk}") # Confirmación
                                except Exception as log_error:
                                    print(f"Error al registrar LogEntry ({'ADD' if action_flag_log == ADDITION else 'CHANGE'} Venta Bomba): {log_error}")
                                # --- FIN OPCIONAL ---

                        num_ventas_guardadas_total += num_saved_this_fs
                        num_ventas_borradas_total += num_deleted_this_fs
                        print(f"  Formset lectura {lectura_id}: {num_saved_this_fs} guardadas/actualizadas, {num_deleted_this_fs} borradas.")

                    # Lógica para finalizar turno
                    if finalizando_turno:
                        print("Finalizando turno...")
                        for lectura in lecturas:
                             print(f"  Calculando final para lectura {lectura.id}...")
                             # Es crucial refrescar la instancia por si se borraron ventas asociadas
                             lectura.refresh_from_db()
                             lectura.calcular_y_guardar_final()
                             print(f"  Final calculado: Contador={lectura.contador_final}, Litros={lectura.litros_vendidos_turno}")

                        reporte.fecha_fin = timezone.now()
                        reporte.esta_abierto = False
                        reporte.save(update_fields=['fecha_fin', 'esta_abierto'])
                        print("Reporte de turno marcado como cerrado.")

                        # --- REGISTRAR ACCIÓN FIN TURNO (YA PRESENTE, ASEGURAR QUE FUNCIONE) ---
                        try:
                           LogEntry.objects.log_action(
                               user_id=request.user.id,
                               content_type_id=ContentType.objects.get_for_model(reporte).pk,
                               object_id=reporte.pk,
                               object_repr=str(reporte),
                               action_flag=CHANGE, # Correcto, es un cambio de estado
                               change_message=f"Turno finalizado. {num_ventas_guardadas_total} ventas guardadas/actualizadas, {num_ventas_borradas_total} borradas."
                           )
                           print(f"LogEntry registrado para Finalización de Turno ID: {reporte.pk}") # Confirmación
                        except Exception as log_error:
                           # **ACCIÓN:** Revisa los logs del servidor si esta acción no aparece.
                           print(f"Error al registrar LogEntry para fin de ReporteTurno: {log_error}")
                        # --- FIN REGISTRO ACCIÓN ---
                # --- FIN BLOQUE try with transaction.atomic ---

                # Mensajes y redirección
                if finalizando_turno:
                    messages.success(request, "Turno finalizado y reporte guardado correctamente.")
                    return redirect('nembus_app:dashboard_trabajador')
                else:
                    messages.success(request, f"Avance guardado: {num_ventas_guardadas_total} ventas guardadas/actualizadas, {num_ventas_borradas_total} borradas.")
                    # Refrescar la misma página para ver cambios y permitir añadir más
                    return redirect('nembus_app:gestionar_turno', reporte_id=reporte.id)

            except Exception as e: # Captura de excepciones generales
                 print(f"!!! EXCEPCIÓN durante transaction.atomic: {e}")
                 messages.error(request, f"Error al guardar o finalizar el turno: {e}")
                 formsets_dict = formsets_procesados # Pasa los formsets con errores al contexto
        else: # Si formsets_validos es False
             messages.error(request, "No se pudo guardar. Revisa los errores en los formularios.") # Mensaje genérico
             formsets_dict = formsets_procesados # Pasa los formsets con errores al contexto

    # Método GET o si hubo errores en POST (rellena el contexto)
    lecturas_con_formsets = []
    # Usar los formsets procesados si hubo error en POST, si no, crear nuevos
    formsets_a_usar = formsets_dict if formsets_dict else {}
    for lectura in lecturas:
        prefix = f'ventas_{lectura.id}'
        # Si hubo un error en POST, usa el formset ya instanciado (con errores)
        if lectura.id in formsets_a_usar:
            formset_para_plantilla = formsets_a_usar[lectura.id]
        else: # Si es GET o el formset fue válido, crea uno nuevo para mostrar
            formset_para_plantilla = VentaIndividualFormSet(instance=lectura, prefix=prefix)
        lecturas_con_formsets.append((lectura, formset_para_plantilla))


    context = {
        'reporte': reporte,
        'lecturas_con_formsets': lecturas_con_formsets
    }
    print(f"--- Fin Depuración gestionar_turno ({request.method}) ---")
    return render(request, 'nembus_app/gestionar_turno.html', context)


# --- VISTAS PARA EL GERENTE ---

@login_required
def dashboard_redirect(request):
    """Redirige al dashboard por defecto (camiones, dia)."""
    if not request.user.is_superuser:
        messages.warning(request, "Accediendo al dashboard de trabajador.")
        return redirect('nembus_app:dashboard_trabajador')
    return redirect('nembus_app:dashboard_gerente', division='camiones', periodo='dia')

# --- FUNCIÓN get_periodo_filter (SIN CAMBIOS) ---
def get_periodo_filter(periodo):
    """Devuelve start_dt, end_dt (datetimes con timezone), título y nombre del período."""
    hoy_aware = timezone.now() # Usar datetime con timezone
    hoy_date = hoy_aware.date() # Obtener solo la fecha

    if periodo == 'semana':
        fecha_inicio_date = hoy_date - timedelta(days=hoy_date.weekday()) # Lunes
        fecha_fin_date = fecha_inicio_date + timedelta(days=7) # Próximo Lunes
        titulo = 'Esta Semana'
    elif periodo == 'mes':
        fecha_inicio_date = hoy_date.replace(day=1) # Primer día del mes
        if hoy_date.month == 12:
            fecha_fin_date = hoy_date.replace(year=hoy_date.year + 1, month=1, day=1)
        else:
            fecha_fin_date = hoy_date.replace(month=hoy_date.month + 1, day=1)
        titulo = 'Este Mes'
    else: # Por defecto, 'dia'
        periodo = 'dia'
        fecha_inicio_date = hoy_date # Hoy
        fecha_fin_date = hoy_date + timedelta(days=1) # Mañana
        titulo = 'Hoy'

    # Crear datetimes ingenuos (sin zona horaria) primero
    naive_start_dt = datetime.combine(fecha_inicio_date, datetime.min.time())
    naive_end_dt = datetime.combine(fecha_fin_date, datetime.min.time())

    # Usar timezone.make_aware() para hacerlos conscientes de la zona horaria actual
    start_dt = timezone.make_aware(naive_start_dt)
    end_dt = timezone.make_aware(naive_end_dt)

    return start_dt, end_dt, titulo, periodo


@login_required
def dashboard_gerente(request, division='camiones', periodo='dia'):
    if not request.user.is_superuser:
        messages.error(request, "Acceso denegado.")
        return redirect('nembus_app:dashboard_trabajador')

    start_dt, end_dt, titulo_periodo, periodo_actual = get_periodo_filter(periodo)

    print(f"\n--- Depurando dashboard_gerente ---")
    print(f"Periodo: {periodo_actual}, División: {division}")
    print(f"Filtrando entre: {start_dt} y {end_dt}")

    context = {
        'division_seleccionada': division,
        'periodo_seleccionado': periodo_actual,
        'titulo_periodo': titulo_periodo
    }

    # Ventas de BOMBAS (Filtradas por fecha de registro)
    ventas_bomba_registradas = RegistroVentaIndividualBomba.objects.filter(
        fecha_registro__gte=start_dt, fecha_registro__lt=end_dt
    ).select_related('lectura_bomba__bomba__punto_de_venta', 'lectura_bomba__reporte_turno__turno') # Añadir turno
    print(f"Ventas Bomba encontradas (antes de agregar): {ventas_bomba_registradas.count()}")

    totales_ventas_bomba = ventas_bomba_registradas.aggregate(
        total_litros=Sum('litros_vendidos'),
        total_ingreso=Sum('ingreso_registro')
    )
    context['total_litros_vendidos_bomba_detalle'] = totales_ventas_bomba['total_litros'] or Decimal('0.00')
    context['total_ingreso_bomba_detalle'] = totales_ventas_bomba['total_ingreso'] or Decimal('0.00')
    print(f"Totales Bomba calculados: Litros={context['total_litros_vendidos_bomba_detalle']}, Ingreso={context['total_ingreso_bomba_detalle']}")

    # Ventas de CAMIONES (Filtradas por fecha_hora)
    reportes_camiones = ReporteVenta.objects.filter(
        fecha_hora__gte=start_dt, fecha_hora__lt=end_dt
    ).select_related('cliente', 'camion', 'trabajador')
    print(f"Reportes Camión encontrados: {reportes_camiones.count()}")

    totales_ventas_camion = reportes_camiones.aggregate(
        total_litros=Sum('litros_vendidos'),
        total_ingreso=Sum('monto_total_clp'),
        total_combustible=Sum('monto_combustible_clp'),
        total_flete=Sum('costo_flete_clp')
    )
    context['litros_vendidos_camiones'] = totales_ventas_camion['total_litros'] or Decimal('0.00')
    context['ingresos_totales_camiones'] = totales_ventas_camion['total_ingreso'] or Decimal('0.00')
    context['ingresos_combustible_camiones'] = totales_ventas_camion['total_combustible'] or Decimal('0.00')
    context['ingresos_flete_camiones'] = totales_ventas_camion['total_flete'] or Decimal('0.00')
    context['numero_viajes_camiones'] = reportes_camiones.count()
    context['promedio_litros_viaje'] = (context['litros_vendidos_camiones'] / context['numero_viajes_camiones']) if context['numero_viajes_camiones'] > 0 else Decimal('0.00')
    print(f"Totales Camión calculados: Litros={context['litros_vendidos_camiones']}, Ingreso={context['ingresos_totales_camiones']}")

    # Lógica Específica por División
    if division == 'camiones':
        context['litros_vendidos'] = context['litros_vendidos_camiones']
        context['ingresos_totales'] = context['ingresos_totales_camiones']

        if periodo_actual == 'dia':
            camiones = Camion.objects.all().order_by('patente')
            for c in camiones:
                litros = c.litros_actuales if c.litros_actuales is not None else Decimal('0.00')
                capacidad = c.capacidad_total if c.capacidad_total is not None and c.capacidad_total > 0 else Decimal('1.00')
                # Asegurar que capacidad nunca sea cero para evitar división por cero
                c.porcentaje_actual = (litros / capacidad) * 100 if capacidad > 0 else 0
            context['camiones'] = camiones
            # Añadir inventario de bombas también para la vista 'hoy'
            context['bombas'] = Bomba.objects.select_related('punto_de_venta').order_by('punto_de_venta__nombre', 'nombre')
            # Gráfico de proporción de ingresos camión (combustible vs flete)
            context['ingresos_combustible_hoy'] = float(context['ingresos_combustible_camiones'])
            context['ingresos_flete_hoy'] = float(context['ingresos_flete_camiones'])
            # Gráfico ventas por hora
            ventas_por_hora = reportes_camiones.annotate(hora=TruncHour('fecha_hora')).values('hora').annotate(total_litros=Sum('litros_vendidos')).order_by('hora')
            context['ventas_hora_labels'] = json.dumps([v['hora'].strftime('%H:%M') for v in ventas_por_hora])
            context['ventas_hora_data'] = json.dumps([float(v['total_litros'] or 0) for v in ventas_por_hora])


        if periodo_actual == 'semana' or periodo_actual == 'mes':
            eficiencia = reportes_camiones.values('camion__patente').annotate(
                num_viajes=Count('id'), total_litros=Sum('litros_vendidos'), total_ingresos=Sum('monto_total_clp')
            ).order_by('-total_ingresos')
            for c in eficiencia:
                litros = c['total_litros'] or Decimal('0.00')
                viajes = c['num_viajes'] or 1
                c['promedio_litros_viaje'] = (litros / viajes) if viajes > 0 else Decimal('0.00')
            context['eficiencia_flota'] = eficiencia

        if periodo_actual == 'mes':
            tendencia_camiones = reportes_camiones.annotate(dia=TruncDay('fecha_hora')).values('dia').annotate(total=Sum('litros_vendidos')).order_by('dia')
            context['tendencia_mes_labels'] = json.dumps([v['dia'].strftime('%d/%m') for v in tendencia_camiones])
            context['tendencia_mes_data'] = json.dumps([float(v['total'] or 0) for v in tendencia_camiones])

    elif division == 'bombas':
        # KPIs para bombas
        context['litros_vendidos'] = context['total_litros_vendidos_bomba_detalle']
        context['ingresos_totales'] = context['total_ingreso_bomba_detalle']
        # Contar turnos que INICIARON en el período
        context['turnos_reportados'] = ReporteTurno.objects.filter(
            fecha_inicio__gte=start_dt, fecha_inicio__lt=end_dt
        ).count()

        datos_por_pdv = []
        puntos_de_venta = PuntoDeVenta.objects.all().order_by('nombre')
        for pdv in puntos_de_venta:
            # Filtrar ventas por punto de venta
            ventas_pdv = ventas_bomba_registradas.filter(lectura_bomba__bomba__punto_de_venta=pdv)
            if not ventas_pdv.exists(): continue # Saltar si no hay ventas para este PDV

            # Agrupar por bomba
            ventas_por_bomba = ventas_pdv.values('lectura_bomba__bomba__nombre').annotate(
                total_litros=Sum('litros_vendidos')
            ).order_by('-total_litros')

            # Agrupar por turno nominal
            ventas_por_turno_nominal = ventas_pdv.values('lectura_bomba__reporte_turno__turno__nombre').annotate(
                total_litros=Sum('litros_vendidos')
            ).order_by('-total_litros')

            datos_por_pdv.append({
                'pdv_nombre': pdv.nombre,
                'bombas_labels': json.dumps([b['lectura_bomba__bomba__nombre'] or 'N/A' for b in ventas_por_bomba]),
                'bombas_data': json.dumps([float(b['total_litros'] or 0) for b in ventas_por_bomba]),
                'turnos_labels': json.dumps([t['lectura_bomba__reporte_turno__turno__nombre'] or 'N/A' for t in ventas_por_turno_nominal]),
                'turnos_data': json.dumps([float(t['total_litros'] or 0) for t in ventas_por_turno_nominal]),
            })
        context['datos_detallados_pdv'] = datos_por_pdv

    elif division == 'relaciones':
        context['litros_vendidos_totales'] = context['litros_vendidos_camiones'] + context['total_litros_vendidos_bomba_detalle']
        context['ingresos_totales'] = context['ingresos_totales_camiones'] + context['total_ingreso_bomba_detalle']
        context['ingresos_desglose_data'] = json.dumps([
            float(context['ingresos_combustible_camiones']),
            float(context['ingresos_flete_camiones']),
            float(context['total_ingreso_bomba_detalle'])
        ])

        # Rendimiento choferes (basado en ventas de camión)
        viajes_por_chofer = reportes_camiones.values('trabajador__username').annotate(
            num_viajes=Count('id'), total_ingresos=Sum('monto_total_clp')
        ).order_by('-num_viajes')
        # Filtrar resultados donde el username es None o vacío (si es posible)
        viajes_por_chofer_clean = [v for v in viajes_por_chofer if v['trabajador__username']]
        context['viajes_chofer_labels'] = json.dumps([v['trabajador__username'] for v in viajes_por_chofer_clean])
        context['viajes_chofer_data'] = json.dumps([v['num_viajes'] for v in viajes_por_chofer_clean])

        # Chofer más rentable
        if viajes_por_chofer_clean:
            chofer_rentable_list = sorted(viajes_por_chofer_clean, key=lambda x: x.get('total_ingresos') or Decimal('0.00'), reverse=True)
            context['chofer_rentable'] = chofer_rentable_list[0] if chofer_rentable_list else None
        else:
            context['chofer_rentable'] = None

        # Comparativa clientes (basado en ventas de camión)
        comparativa_clientes = reportes_camiones.values('cliente__nombre').annotate(
            total_litros=Sum('litros_vendidos'),
            total_monto=Sum('monto_total_clp')
        ).order_by('-total_monto')[:5] # Top 5
        context['comparativa_clientes_labels'] = json.dumps([c['cliente__nombre'] or 'N/A' for c in comparativa_clientes])
        context['comparativa_clientes_litros'] = json.dumps([float(c['total_litros'] or 0) for c in comparativa_clientes])
        context['comparativa_clientes_monto'] = json.dumps([float(c['total_monto'] or 0) for c in comparativa_clientes])

    print(f"--- Fin Depuración dashboard_gerente ---")
    return render(request, 'nembus_app/dashboard_gerente.html', context)


# --- VISTAS DE EXPORTACIÓN ---

@login_required
def exportar_reportes_csv(request): # EXPORTACIÓN CSV - SOLO PARA CAMIONES
    if not request.user.is_superuser:
        messages.error(request, "Acceso denegado.")
        return redirect('nembus_app:dashboard_trabajador')

    periodo_seleccionado = request.GET.get('periodo', 'dia')
    start_dt, end_dt, _, _ = get_periodo_filter(periodo_seleccionado)

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response.write(u'\ufeff'.encode('utf8')) # BOM para Excel
    writer = csv.writer(response, delimiter=';') # Usar punto y coma

    reportes = ReporteVenta.objects.filter(
        fecha_hora__gte=start_dt, fecha_hora__lt=end_dt
    ).select_related('trabajador', 'cliente', 'camion').order_by('fecha_hora')

    filename = f'reporte_ventas_camiones_{periodo_seleccionado}_{timezone.now().strftime("%Y%m%d")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"' # Comillas por si acaso

    writer.writerow(['Fecha', 'Hora', 'Trabajador', 'Cliente', 'Camion', 'Litros Vendidos', 'Monto Combustible (CLP)', 'Costo Flete (CLP)', 'Monto Total (CLP)'])

    for reporte in reportes:
        fecha_hora_local = timezone.localtime(reporte.fecha_hora)
        writer.writerow([
            fecha_hora_local.strftime('%Y-%m-%d'),
            fecha_hora_local.strftime('%H:%M:%S'),
            reporte.trabajador.username if reporte.trabajador else 'N/A',
            reporte.cliente.nombre if reporte.cliente else 'N/A',
            reporte.camion.patente if reporte.camion else 'N/A',
            # Usar punto como separador decimal para CSV estándar, Excel debería reconocerlo
            str(reporte.litros_vendidos).replace(',', '.'),
            str(reporte.monto_combustible_clp).replace(',', '.'),
            str(reporte.costo_flete_clp).replace(',', '.'),
            str(reporte.monto_total_clp).replace(',', '.')
        ])

    return response


@login_required
def exportar_ventas_bomba_excel(request):
    if not request.user.is_superuser:
        messages.error(request, "Acceso denegado.")
        return redirect('nembus_app:dashboard_trabajador')

    print(f"\n--- Depurando exportar_ventas_bomba_excel ---")
    periodo = request.GET.get('periodo', 'todos')
    # --- NUEVO: Obtener el ID del Punto de Venta ---
    punto_venta_id = request.GET.get('punto_venta_id', None) # Obtiene el ID, default None
    punto_venta_seleccionado = None # Para guardar el objeto o nombre
    print(f"Periodo solicitado: {periodo}")
    print(f"Punto de Venta ID solicitado: {punto_venta_id}") # Para depurar

    # Query inicial
    ventas_query = RegistroVentaIndividualBomba.objects.select_related(
        'lectura_bomba__bomba__punto_de_venta',
        'lectura_bomba__reporte_turno__trabajador',
        'lectura_bomba__reporte_turno__turno'
    ).all() # Empezamos con .all()

    # --- NUEVO: Aplicar filtro por Punto de Venta si se proporcionó un ID ---
    if punto_venta_id and punto_venta_id.isdigit(): # Verifica que sea un ID numérico válido
        try:
            # Filtra el queryset principal
            ventas_query = ventas_query.filter(lectura_bomba__bomba__punto_de_venta_id=int(punto_venta_id))
            # Opcional: Obtener el nombre para mostrarlo en el Excel
            punto_venta_seleccionado = PuntoDeVenta.objects.get(id=int(punto_venta_id))
            print(f"Filtrando por Punto de Venta: {punto_venta_seleccionado.nombre}")
        except PuntoDeVenta.DoesNotExist:
            messages.error(request, "Punto de venta no encontrado.")
            # Decide si retornar un error o exportar todo
            print(f"Punto de Venta ID {punto_venta_id} no encontrado. Exportando todo.")
            punto_venta_id = None # Anula el ID para no usarlo en el nombre de archivo
            punto_venta_seleccionado = None
    else:
        print("No se especificó Punto de Venta ID o no es válido. Exportando todos.")
        punto_venta_id = None # Asegura que no se use un ID inválido

    # Aplicar filtro por período (después del filtro de punto de venta)
    periodo_filtro = periodo if periodo != 'todos' else 'dia'
    start_dt, end_dt, _, _ = get_periodo_filter(periodo_filtro)

    if periodo != 'todos':
        print(f"Filtrando Excel entre: {start_dt} y {end_dt}")
        if isinstance(start_dt, datetime):
            ventas_query = ventas_query.filter(fecha_registro__gte=start_dt, fecha_registro__lt=end_dt)
        else:
            start_dt_aware = timezone.make_aware(datetime.combine(start_dt, datetime.min.time()))
            end_dt_aware = timezone.make_aware(datetime.combine(end_dt, datetime.min.time()))
            ventas_query = ventas_query.filter(fecha_registro__gte=start_dt_aware, fecha_registro__lt=end_dt_aware)
    else:
        print("Exportando todos los registros (periodo=todos)")

    # Ordenar al final, después de todos los filtros
    ventas_query = ventas_query.order_by('fecha_registro')

    print(f"Ventas encontradas para Excel: {ventas_query.count()}")

    # --- Creación del Excel ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # --- NUEVO: Añadir nombre del punto de venta al filename si se filtró ---
    pv_suffix = f"_{punto_venta_seleccionado.nombre.replace(' ','_')}" if punto_venta_seleccionado else ""
    filename = f'reporte_ventas_bombas{pv_suffix}_{periodo}_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Ventas desde Bombas"

    # --- Estilos (sin cambios) ---
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    currency_format = '$ #,##0'
    litros_format = '#,##0.00 "L"'
    price_format = '$ #,##0.00'

    # --- Cabecera del Reporte ---
    sheet.merge_cells('A1:J1')
    cell_A1 = sheet['A1']
    # --- NUEVO: Incluir nombre del PV en el título si se filtró ---
    titulo_reporte = "INFORME DE VENTAS DESDE BOMBAS"
    if punto_venta_seleccionado:
        titulo_reporte += f" - {punto_venta_seleccionado.nombre}"
    cell_A1.value = titulo_reporte
    cell_A1.font = Font(bold=True, size=16)
    cell_A1.alignment = align_center

    hora_local_santiago = timezone.localtime(timezone.now())
    sheet['A3'] = "Fecha de Reporte:"
    sheet['B3'] = hora_local_santiago.strftime("%d/%m/%Y %H:%M:%S")

    # --- Lógica para PERIODO en celda B4 (modificada en paso anterior) ---
    sheet['A4'] = "Período:"
    periodo_texto = ""
    # (Aquí va la lógica if/elif/else para periodo_texto que definimos antes)
    if periodo == 'todos':
        periodo_texto = 'Todos los registros'
    elif periodo == 'dia':
        # ... (código para formatear día) ...
         if isinstance(start_dt, datetime): periodo_texto = start_dt.strftime("%d/%m/%Y")
         else: periodo_texto = start_dt.strftime("%d/%m/%Y")
    elif periodo == 'semana':
        # ... (código para formatear semana) ...
        fecha_fin_semana = end_dt - timedelta(days=1)
        if isinstance(start_dt, datetime):
            fecha_inicio_str = start_dt.strftime("%d/%m/%Y"); fecha_fin_str = fecha_fin_semana.strftime("%d/%m/%Y")
        else:
            fecha_inicio_str = start_dt.strftime("%d/%m/%Y"); fecha_fin_str = fecha_fin_semana.strftime("%d/%m/%Y")
        periodo_texto = f"Semana del {fecha_inicio_str} al {fecha_fin_str}"
    elif periodo == 'mes':
        # ... (código para formatear mes con locale) ...
        try: locale.setlocale(locale.LC_TIME, 'es_CL.UTF-8')
        except locale.Error:
            try: locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
            except locale.Error: locale.setlocale(locale.LC_TIME, '')
        if isinstance(start_dt, datetime): nombre_mes = start_dt.strftime("%B").capitalize(); anio = start_dt.year
        else: nombre_mes = start_dt.strftime("%B").capitalize(); anio = start_dt.year
        periodo_texto = f"{nombre_mes} {anio}"
    else:
        periodo_texto = periodo.capitalize()
    sheet['B4'] = periodo_texto

    # --- NUEVO: Añadir fila para Punto de Venta si se filtró ---
    if punto_venta_seleccionado:
        sheet['A5'] = "Punto de Venta:"
        sheet['B5'] = punto_venta_seleccionado.nombre
        sheet['A5'].font = Font(bold=True) # Opcional: poner en negrita

    # --- Encabezados de Tabla ---
    headers = [
        "Turno", "Fecha Pago", "Máquina", "Socio", "Pagado (CLP)",
        "Litros", "Precio Litro", "Bomba", "Trabajador", "Punto Venta"
    ]
    sheet.append([]) # Fila vacía
    # Ajusta el número de fila inicial para los encabezados y datos
    header_row_num = sheet.max_row + 1 # Ahora será 7 si se añadió la fila del PV

    # --- Asignación celda por celda (sin cambios en la lógica) ---
    for col_num, header_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell_address = f"{col_letter}{header_row_num}"
        sheet[cell_address] = header_title

    # --- Aplicar Estilo a Encabezados (sin cambios en la lógica) ---
    for col_num, cell in enumerate(sheet[header_row_num], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_style
    sheet.row_dimensions[header_row_num].height = 30

    # --- Datos ---
    total_litros = Decimal('0.00')
    total_ingreso = Decimal('0.00')
    row_num = header_row_num # Inicializa para empezar DESPUÉS de los encabezados

    for venta in ventas_query:
        row_num += 1
        # ... (código para obtener turno_nombre, bomba_nombre, etc.) ...
        turno_nombre = venta.lectura_bomba.reporte_turno.turno.nombre if venta.lectura_bomba and venta.lectura_bomba.reporte_turno and venta.lectura_bomba.reporte_turno.turno else 'N/A'
        bomba_nombre = venta.lectura_bomba.bomba.nombre if venta.lectura_bomba and venta.lectura_bomba.bomba else 'N/A'
        trabajador_nombre = venta.lectura_bomba.reporte_turno.trabajador.username if venta.lectura_bomba and venta.lectura_bomba.reporte_turno and venta.lectura_bomba.reporte_turno.trabajador else 'N/A'
        pdv_nombre = venta.lectura_bomba.bomba.punto_de_venta.nombre if venta.lectura_bomba and venta.lectura_bomba.bomba and venta.lectura_bomba.bomba.punto_de_venta else 'N/A'

        # Asignación de datos a celdas (sin cambios en la lógica)
        sheet.cell(row=row_num, column=1, value=turno_nombre)
        fecha_pago = venta.fecha_registro.strftime("%d/%m/%Y") if isinstance(venta.fecha_registro, datetime) else 'N/A'
        sheet.cell(row=row_num, column=2, value=fecha_pago)
        sheet.cell(row=row_num, column=3, value=getattr(venta, 'numero_maquina', 'N/A'))
        sheet.cell(row=row_num, column=4, value=getattr(venta, 'socio_propietario', 'N/A'))
        sheet.cell(row=row_num, column=5, value=getattr(venta, 'ingreso_registro', Decimal('0.00')))
        sheet.cell(row=row_num, column=6, value=getattr(venta, 'litros_vendidos', Decimal('0.00')))
        sheet.cell(row=row_num, column=7, value=getattr(venta, 'precio_litro_venta', Decimal('0.00')))
        sheet.cell(row=row_num, column=8, value=bomba_nombre)
        sheet.cell(row=row_num, column=9, value=trabajador_nombre)
        sheet.cell(row=row_num, column=10, value=pdv_nombre)

        # ... (código para sumar totales y aplicar formatos/bordes a las celdas de datos) ...
        total_litros += venta.litros_vendidos if venta.litros_vendidos else Decimal('0.00')
        total_ingreso += venta.ingreso_registro if venta.ingreso_registro else Decimal('0.00')
        sheet.cell(row=row_num, column=5).number_format = currency_format
        sheet.cell(row=row_num, column=6).number_format = litros_format
        precio_litro_cell = sheet.cell(row=row_num, column=7)
        precio_litro_cell.number_format = price_format
        precio_litro_cell.value = venta.precio_litro_venta if venta.precio_litro_venta else Decimal('0.00')
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_num, column=col_idx)
            cell.border = border_style
            if col_idx in [5, 6, 7]: cell.alignment = align_right


    # --- Totales (sin cambios en la lógica) ---
    sheet.append([])
    total_row_num = sheet.max_row + 1
    cell_total_label = sheet.cell(row=total_row_num, column=4, value="TOTALES:")
    cell_total_label.font = Font(bold=True)
    cell_total_label.alignment = align_right
    cell_total_label.border = border_style
    # ... (código para celdas de total_ingreso y total_litros y bordes de celdas vacías) ...
    cell_total_ingreso = sheet.cell(row=total_row_num, column=5, value=total_ingreso)
    cell_total_ingreso.number_format = currency_format
    cell_total_ingreso.font = Font(bold=True)
    cell_total_ingreso.border = border_style
    cell_total_ingreso.alignment = align_right
    cell_total_litros = sheet.cell(row=total_row_num, column=6, value=total_litros)
    cell_total_litros.number_format = litros_format
    cell_total_litros.font = Font(bold=True)
    cell_total_litros.border = border_style
    cell_total_litros.alignment = align_right
    for col in [1, 2, 3, 7, 8, 9, 10]: sheet.cell(row=total_row_num, column=col).border = border_style


    # --- Ajustar Ancho de Columnas (sin cambios) ---
    column_widths = {'A': 12, 'B': 12, 'C': 15, 'D': 25, 'E': 15, 'F': 12, 'G': 12, 'H': 20, 'I': 15, 'J': 20}
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # --- Guardar y devolver ---
    workbook.save(response)
    print(f"--- Fin Depuración Excel ---")
    return response

# --- Función get_periodo_filter (incluida para completitud) ---
def get_periodo_filter(periodo): #
    # ... (código de get_periodo_filter) ...
    hoy = timezone.now().date()
    if periodo == 'semana':
        fecha_inicio = hoy - timedelta(days=hoy.weekday())
        fecha_fin = fecha_inicio + timedelta(days=7)
        titulo = 'Esta Semana'
    elif periodo == 'mes':
        fecha_inicio = hoy.replace(day=1)
        if hoy.month == 12: fecha_fin = hoy.replace(year=hoy.year + 1, month=1, day=1)
        else: fecha_fin = hoy.replace(month=hoy.month + 1, day=1)
        titulo = 'Este Mes'
    else: # dia
        periodo = 'dia'; fecha_inicio = hoy; fecha_fin = hoy + timedelta(days=1); titulo = 'Hoy'
    return fecha_inicio, fecha_fin, titulo, periodo